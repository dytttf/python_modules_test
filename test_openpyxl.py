#coding:utf8
'''
openpyxl 模块
支持同时读写excel文件
测试代码
'''
import os
import openpyxl


cur_path = os.path.dirname(os.path.abspath(__file__))
debug_file = os.path.join(cur_path, "test_files", "test_openpyxl.xlsx")
debug_change_file = os.path.join(cur_path, "test_files", "test_openpyxl_change.xlsx")

# 创建测试文件代码
if not os.path.exists(debug_file):
    print(u"不存在测试文件 正在创建...")
    # 声明一个excel文件对象
    wb = openpyxl.Workbook()
    # Sheet 表为默认存在的第一个表
    sheet = wb.get_sheet_by_name("Sheet")
    # 修改表名
    sheet.title = "Test Sheet 1"
    # 写入数据
    data = {
        "A1": u"姓名",
        "A2": "Alice",
        "A3": "Bob",
        "B1": u"年龄",
        "B2": "21",
        "B3": "34",
        }
    for position, value in data.iteritems():
        sheet[position] = value
        # sheet.cell(row=1, column=1, value=value)
    wb.save(debug_file)
    print(u"创建成功\n {} ".format(debug_file))

# 修改文件测试 另存为新的文件

# 加载文件
print(u"加载测试文件 请注意大文件加载时间会比较长")
wb = openpyxl.load_workbook(debug_file)

# 获取表名列表
sheet_names = wb.get_sheet_names()

# 获取操作表
work_sheet = wb.get_sheet_by_name("Test Sheet 1")

# 获取行数和列数
print(u"行数: {}".format(work_sheet.max_row))
print(u"列数: {}".format(work_sheet.max_column))

# 获取单元格数据
A1 = work_sheet['A1']
print(u"单元格 A1 数据为: {}".format(A1.value))
# 支持 A1:B3 这样的格式
A1_B3 = work_sheet['A1:B3']
# (
#    (<Cell Test Sheet 1.A1>, <Cell Test Sheet 1.B1>),
#    (<Cell Test Sheet 1.A2>, <Cell Test Sheet 1.B2>),
#    (<Cell Test Sheet 1.A3>, <Cell Test Sheet 1.B3>)
# )

# 修改数据
A1.value = "name"
print(u"修改后 单元格 A1 数据为: {}".format(A1.value))

# 不知道怎么复制一个 sheet ...

# 保存修改到新的表
wb.save(debug_change_file)



import MySQLdb
import openpyxl

def main():
##    mysql_conn = MySQLdb.connect(
##        host='localhost',
##        port=3306,
##        user='root',
##        passwd='123456',
##        db='test',
##        charset='utf8'
##        )
##    # 字典
##    cursor = mysql_conn.cursor(cursorclass=MySQLdb.cursors.DictCursor)
##    sql = ""
##
##    # 执行sql 将数据获取到内存中
##    cursor.execute(sql)
##    datas = cursor.fetchall()
    datas = [
        {
        "A":'1',
        "B":'2',
            },
        {
        "A":'3',
        "B":'4',
            }
        ]
    # 新建excel
    wb = openpyxl.Workbook()
    sheet = wb.get_sheet_by_name("Sheet")
    # 修改表名
    sheet.title = "TestSheet1"

    # 写入表头
    sort_keys = datas[0].keys()
    sort_keys.sort()
    for index, key in enumerate(sort_keys):
        sheet.cell(row=1, column=index+1, value=key)
    # 写入数据
    for line, data in enumerate(datas):
        for index, key in enumerate(sort_keys):
            sheet.cell(row=line+2, column=index+1, value=str(data[key]))
    wb.save('test.xlsx')
    return

    










